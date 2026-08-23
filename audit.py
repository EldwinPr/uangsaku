#!/usr/bin/env python3
"""Consistency audit for this repo's documentation. Run: python audit.py

Checks that the artifacts agree with each other. It cannot check that any of them is
RIGHT, only that they do not contradict each other or point at things that do not
exist. A clean run means the paper trail is intact, not that the design is good.

Requires: openpyxl, pyyaml.
"""
import os
import re
import glob
import sys
import hashlib
import xml.dom.minidom

import openpyxl
import yaml

fail, warn, ok = [], [], []
F, W, O = fail.append, warn.append, ok.append

# Paths a document may legitimately name even though the file is gone or was never
# real. Each carries its reason -- do not add entries here to silence a real check.
ALLOWED_DANGLING = {
    'docs/diagrams/state-budget-period.drawio':
        'deleted by ISSUE-004; cited only by the documents recording the deletion',
    'docs/diagrams/class-budgeting-draft.drawio':
        'deleted 2026-08-20 as superseded; cited only where that is explained',
    'docs/diagrams/bpmn-as-is.drawio':
        'generic filename in bpmn-conventions.md, not a claim that a file exists',
    'docs/diagrams/bpmn-to-be.drawio':
        'same as bpmn-as-is.drawio -- this project uses the FR route, so has no BPMN',
    'pm/issues/015-sequence-conventions/plan.md':
        'renumbered to 008-sequence-conventions on 2026-08-20; pm/log.md is append-only, '
        'so entries written before the renumber still name the old path',
    'docs/rbac-entities.csv':
        'deleted 2026-08-20 at the owner request (empty, and a single-actor app has no '
        'roles); cited only by the records of that deletion and by ISSUE-006 which '
        'replaces it with docs/enums.md',
}

# Files a document names because they are PLANNED, not because they exist. Distinct
# from ALLOWED_DANGLING (things deleted or never real): these are expected to appear.
# The plan.md of every TODO tracker issue is added to this set automatically below.
PLANNED_NOT_WRITTEN = {
    # Empty, and that is the intended resting state: an entry here is a promise that a
    # named file will exist, not a permanent exemption. sequence-conventions.md left this
    # set on 2026-08-20 (ISSUE-008), context/coding-conventions/ on the same day
    # (ISSUE-009). The plan.md of every TODO tracker issue is added below automatically.
}

# The plan of a TODO issue does not exist yet by definition -- see the tracker check.
try:
    _tr = yaml.safe_load(open('pm/tracker.yaml', encoding='utf-8'))['issues']
    PLANNED_NOT_WRITTEN.update(
        {i['plan']: 'plan of {}, still TODO'.format(i['id'])
         for i in _tr if i.get('status') == 'TODO'})
except Exception as _e:  # the tracker's own checks report this properly further down
    fail.append('could not read pm/tracker.yaml for the planned-file set: {}'.format(_e))

def _sha(path):
    """Hash of a diagram source, used to tell a stale render from a current one.

    Line endings are normalised first, and that is not a detail. This repo is developed
    on Windows with core.autocrlf=true, so a checkout rewrites LF to CRLF in the working
    tree while git stores LF -- a raw byte hash therefore changes on clone, and every
    render would read as stale on CI and fresh locally, or the reverse. Normalising
    means the hash describes the diagram rather than the machine it was checked out on.
    """
    # newline=None is universal-newline mode: CRLF and LF both read back as LF, with
    # no escape sequences needed to say so.
    text = open(path, encoding='utf-8', newline=None).read()
    return hashlib.sha256(text.encode('utf-8')).hexdigest()


def _slugdir(issue_id):
    """pm/issues/ folder name for a tracker id -- the tracker does not store the path.

    Matches on the exact first hyphen-separated token, not a prefix: `startswith`
    let `UC02-add-account` match either `uc02-add-account` or `uc02b-edit-account`
    (both satisfy `.startswith('uc02')`), and `glob.glob`'s return order is
    filesystem-dependent -- stable enough on this project's Windows dev machine to
    pick the right one by luck, but not guaranteed on Linux (CI), which sometimes
    picked the wrong directory and broke a render-path lookup that passed locally
    every time. `sorted()` also makes the fallback loop itself deterministic.
    """
    prefix = issue_id.lower().split('-')[0]
    for d in sorted(glob.glob('pm/issues/*')):
        base = os.path.basename(d)
        if base.split('-')[0] == prefix:
            return base
    return issue_id.lower()


if '--record-renders' in sys.argv:
    # Run this immediately AFTER re-exporting a sequence diagram's PNG. It records the
    # hash of each .drawio as it stands now, which is the claim "the committed render was
    # made from this exact source". Running it without re-exporting launders a stale
    # picture into looking current, which is the one thing this mechanism exists to catch.
    import yaml as _y
    _lock = {p_.replace(os.sep, '/'): {'source_sha': _sha(p_),
                                       'render': os.path.basename(p_)[:-7] + '.png'}
             for p_ in sorted(glob.glob('docs/diagrams/seq-uc*.drawio'))}
    _header = [
        '# Written by `python audit.py --record-renders`, immediately after a sequence',
        '# diagram is re-exported. Maps each .drawio to the hash of the source its',
        '# committed PNG was made from; audit.py fails when the two drift apart.',
        "# Sequence renders are committed and carry an issue's scope, so a wrong picture",
        '# is worse than a missing one. See drawio-general-guide.md.',
    ]
    with open('docs/diagrams/renders.lock', 'w', encoding='utf-8') as fh:
        fh.write('\n'.join(_header) + '\n')
        _y.safe_dump(_lock, fh, default_flow_style=False, sort_keys=True)
    print('recorded {} sequence renders in docs/diagrams/renders.lock'.format(len(_lock)))
    sys.exit(0)


def is_stub(path):
    """True for an issue folder's NOT PLANNED placeholder, which is not a plan."""
    try:
        return '**Status:** NOT PLANNED' in open(path, encoding='utf-8').read(400)
    except OSError:
        return False


# ---------- diagrams parse, and carry no XML comments ----------
proj = glob.glob('docs/diagrams/*.drawio')
for f in proj + glob.glob('context/**/*.drawio', recursive=True):
    try:
        xml.dom.minidom.parse(f)
    except Exception as e:
        F('XML malformed: {} -- {}'.format(f, e))
    if open(f, encoding='utf-8').read().count('<!--'):
        F('XML comment present (forbidden by drawio-general-guide.md): {}'.format(f))
O('{} project diagrams parse, 0 XML comments'.format(len(proj)))

# ---------- every referenced file exists ----------
gitignored = re.compile(r'docs/diagrams/.*\.png$|context/files/.*\.pdf$')
docs = [p for p in glob.glob('**/*.md', recursive=True)
        + glob.glob('**/*.yaml', recursive=True) if '.claude' not in p]
refs = {}
for p in docs:
    body = open(p, encoding='utf-8').read()
    pattern = (r'(?:docs|pm|context|input)/[A-Za-z0-9_./-]+'
               r'\.(?:md|yaml|xlsx|drawio|csv|png|pdf|txt|json)')
    for m in re.findall(pattern, body):
        refs.setdefault(m.replace('\\', '/'), set()).add(p.replace('\\', '/'))
for r, srcs in sorted(refs.items()):
    if (os.path.exists(r) or gitignored.search(r) or r in ALLOWED_DANGLING
            or r in PLANNED_NOT_WRITTEN):
        continue
    F('referenced but missing: {}  <- cited in {}'.format(r, ', '.join(sorted(srcs))))
# A planned file that HAS been written should leave the planned set, or the set rots
# into a permanent excuse list.
for r, why in sorted(PLANNED_NOT_WRITTEN.items()):
    if os.path.exists(r) and not r.startswith('pm/issues/'):
        W('{} now exists -- drop it from PLANNED_NOT_WRITTEN ({})'.format(r, why))
# Symmetric check, and the reason this one exists: seq-uc02-add-account.drawio sat in
# ALLOWED_DANGLING as "an illustrative filename, not a claim that the file exists" long
# after the file was actually drawn. An exemption that has quietly stopped applying is
# the same defect as one that never applied -- it stops a real check without saying so.
for r, why in sorted(ALLOWED_DANGLING.items()):
    if os.path.exists(r):
        W('{} exists now -- drop it from ALLOWED_DANGLING ({})'.format(r, why))
O('{} distinct file references checked ({} known-dangling, {} planned-not-written)'.format(
    len(refs), len(ALLOWED_DANGLING), len(PLANNED_NOT_WRITTEN)))

# ---------- a gitignored-but-cited spec must leave a stub ----------
for r in refs:
    if r.startswith('context/files/') and r.endswith('.pdf') and not os.path.exists(r):
        if not os.path.exists(r + '.txt'):
            F('cited spec has neither the file nor a stub: {} '
              '(CLAUDE.md citation gate)'.format(r))
O('cited external specs all resolve to a file or a stub')

# ---------- workbook internal consistency ----------
wb = openpyxl.load_workbook('docs/workbook.xlsx')


def rows(name):
    ws = wb[name]
    head = [c.value for c in ws[1]]
    return [dict(zip(head, r)) for r in ws.iter_rows(min_row=2, values_only=True)
            if any(v is not None for v in r)]


fr, nonfr = rows('UC FR'), rows('UC Non-FR')
mods = {r['Modul'] for r in rows('Modules')}
ents = {r['Entity yang dibutuhkan'] for r in rows('Entities')}
codes = [r['Kode'] for r in fr + nonfr]
dup = sorted({c for c in codes if codes.count(c) > 1})
if dup:
    F('duplicate Kode across the UC sheets: {}'.format(dup))
else:
    O('{} UC codes, all unique ({}..{})'.format(len(codes), min(codes), max(codes)))
required = ['Kode', 'Nama Use Case', 'User', 'Modul', 'Input', 'Deskripsi', 'Output',
            'Entity/Objek Terkait']
for r in fr + nonfr:
    for col in required:
        if not r.get(col):
            F('{}: empty required column "{}"'.format(r['Kode'], col))
    if r['Modul'] not in mods:
        F('{}: Modul "{}" is not on the Modules sheet'.format(r['Kode'], r['Modul']))
    for e in [x.strip() for x in str(r['Entity/Objek Terkait']).split(',')]:
        if e and e not in ents:
            F('{}: entity "{}" is not on the Entities sheet'.format(r['Kode'], e))
O('workbook: {} UC FR rows, {} UC Non-FR rows, modules and entities all resolve'.format(
    len(fr), len(nonfr)))

# ---------- entities agree across ERD, workbook, map.yaml ----------
erd = open('docs/diagrams/erd.drawio', encoding='utf-8').read()
absent = {e for e in ents if not re.search(r'value="[^"]*' + re.escape(e), erd)}
if absent:
    F('on the Entities sheet but not found on the ERD: {}'.format(sorted(absent)))
m = yaml.safe_load(open('context/index/map.yaml', encoding='utf-8'))
if set(m['entities']) != ents:
    F('map.yaml entities disagree with the workbook: {}'.format(
        sorted(set(m['entities']) ^ ents)))
else:
    O('{} entities agree across the ERD, the workbook and map.yaml'.format(len(ents)))

# ---------- every UC reaches a class diagram, in its own module ----------
covered = set()
for v in m['class_diagrams'].values():
    covered |= set(v['ucs'])
allucs = set(codes)
if allucs - covered:
    F('UC in the workbook but on no class diagram: {}'.format(sorted(allucs - covered)))
if covered - allucs:
    F('UC on a class diagram but not in the workbook: {}'.format(sorted(covered - allucs)))
if not allucs ^ covered:
    O('all {} UCs map to a class diagram'.format(len(allucs)))
mod_of = {r['Kode']: r['Modul'] for r in fr + nonfr}
for mod, v in m['class_diagrams'].items():
    bad = [u for u in v['ucs'] if mod_of.get(u) != mod]
    if bad:
        F('class diagram {} claims UCs whose workbook Modul differs: {}'.format(mod, bad))

# ---------- FR traceability closure ----------
s = open('docs/fr-nfr.md', encoding='utf-8').read()
# Derive the FR set from the document rather than hard-coding a range: a hard-coded
# 1..18 silently stopped checking the moment FR-19 was added (ISSUE-007).
defined = set(re.findall(r'\*\*(FR-\d+) —', s))
if not defined:
    F('no FR definitions found in fr-nfr.md -- the definition pattern may have changed')
top = max(int(x.split('-')[1]) for x in defined) if defined else 0
declared = {'FR-{}'.format(i) for i in range(1, top + 1)}
if declared - defined:
    F('gap in the FR numbering, defined up to FR-{}: {}'.format(
        top, sorted(declared - defined)))
trace = s.split('## 5. Traceability')[1]
traced = {r.split('|')[1].strip() for r in trace.split('\n') if r.startswith('| FR-')}
if declared - traced:
    F('FR defined but missing from the section 5 traceability table: {}'.format(
        sorted(declared - traced)))
else:
    O('all {} FRs are defined and traced'.format(len(declared)))
cited = set(re.findall(r'UC-\d\d', trace))
if cited - allucs:
    F('section 5 cites a UC that is not in the workbook: {}'.format(sorted(cited - allucs)))
if allucs - cited:
    W('UC in the workbook that no FR traces to: {}'.format(sorted(allucs - cited)))

# ---------- tracker / plan consistency ----------
tr = yaml.safe_load(open('pm/tracker.yaml', encoding='utf-8'))['issues']
for i in tr:
    if not os.path.exists(i['plan']):
        # A TODO issue is a backlog entry that has not been planned yet; the gate in
        # general-rules.md requires a plan before work starts, not before the issue is
        # written down. Anything past TODO without a plan is a real violation.
        if i['status'] != 'TODO':
            F('{} is {} but has no plan.md at {}'.format(i['id'], i['status'], i['plan']))
        continue
    # A placeholder is not a plan and must never satisfy the planning gate.
    if is_stub(i['plan']) and i['status'] != 'TODO':
        F('{} is {} but its plan.md is still the NOT PLANNED placeholder'.format(
            i['id'], i['status']))
    for d in i['depends_on']:
        dep = [x for x in tr if x['id'] == d]
        if not dep:
            F('{}: depends on unknown issue {}'.format(i['id'], d))
        elif i['status'] != 'TODO' and dep[0]['status'] != 'DONE':
            F('{} is {} but its dependency {} is {}'.format(
                i['id'], i['status'], d, dep[0]['status']))
    if not i.get('summary'):
        F('{}: no summary on the tracker row'.format(i['id']))
    head = open(i['plan'], encoding='utf-8').read()[:400]
    if i['status'] == 'DONE' and not re.search(
            r'(?i)\*\*status:\*\*\s*(done|superseded)', head):
        W('{}: tracker says DONE but the plan.md status line does not'.format(i['id']))
planned = sum(1 for i in tr if os.path.exists(i['plan']) and not is_stub(i['plan']))
# ---------- every use case is owned by exactly one primary issue ----------
# The backlog is UC-coded as of 2026-08-20, which is a claim worth enforcing: a use
# case with no issue is unbuilt work nobody has noticed, and one claimed twice is two
# issues editing the same screen. Since the Q2 ruling of 2026-08-22 a UC may carry a
# second issue with a letter suffix on its id (UC02B-edit-account -> UC-02): one
# PRIMARY issue per UC keeps the 1:1 claim, and any number of suffixed ones may share
# it -- but never two primaries.
def _issue_uc(issue_id):
    m = re.match(r'UC(\d+)', issue_id.split('-')[0])
    return 'UC-' + m.group(1) if m else None

def _is_primary(issue_id):
    return re.fullmatch(r'UC\d+-[a-z0-9-]+', issue_id) is not None

owner_of = {}
for i in tr:
    for uc in i.get('traces_to') or []:
        owner_of.setdefault(uc, []).append(i['id'])
impl = [i for i in tr if i['id'][:2] in ('UC', 'FE')]
impl_ucs = {}
for i in impl:
    for uc in i.get('traces_to') or []:
        impl_ucs.setdefault(uc, []).append(i['id'])
for uc in sorted(allucs - set(impl_ucs)):
    F('{} is in the workbook but no implementation issue claims it'.format(uc))
splits = 0
for uc, owners in sorted(impl_ucs.items()):
    if uc not in allucs:
        F('issue(s) {} claim {}, which is not a workbook UC'.format(owners, uc))
        continue
    primaries = [x for x in owners if _is_primary(x)]
    if len(primaries) > 1:
        F('{} is claimed by more than one primary implementation issue: {}'.format(uc, owners))
    else:
        if len(owners) > 1:
            splits += 1
            O('{} is shared by {} (primary) and {} (Q2 ruling 2026-08-22: '
              'alternate flows split out under the B-suffix scheme)'.format(
                  uc, primaries[0], ', '.join(x for x in owners if x != primaries[0])))
# An implementation issue's id should match the UC it traces to, or be a FEAT.
# A letter suffix on the id's UC code (UC02B -> UC-02) is the Q2-ruling scheme, not a
# mismatch.
for i in impl:
    if i['id'].startswith('FEAT'):
        continue
    expect = _issue_uc(i['id'])
    if expect is None or expect not in (i.get('traces_to') or []):
        F('{}: id implies {} but traces_to is {}'.format(
            i['id'], expect or 'no UC', i.get('traces_to')))
O('all {} use cases owned by exactly one primary of {} implementation issues '
  '({} shared under the B-suffix scheme); ids match their use case'.format(
      len(allucs), len(impl), splits))

O('{} tracker issues ({} planned, {} still TODO and unplanned by design): '
  'dependencies satisfied, summaries present'.format(
      len(tr), planned, len(tr) - planned))

# ---------- sequence renders: present, in the right issue, and not stale ----------
# A sequence diagram's picture IS its issue's scope (CLAUDE.md), and unlike the other
# diagram types its render is committed. So a .drawio edited without re-exporting leaves
# a wrong picture in the repo where a reader will believe it. Neither git nor a mtime
# check can see that -- mtimes do not survive a clone -- so the export records the hash
# of the source it was made from, and this compares.
RENDER_LOCK = 'docs/diagrams/renders.lock'
# A UC shared under the B-suffix scheme keeps its sequence render with its PRIMARY
# issue for as long as the suffixed one has drawn no diagram of its own. Once a
# suffixed issue draws its own seq-uc{NN}{letter}-*.drawio (UC02B did, 2026-08-23),
# that file names its own issue directly rather than falling back to the primary's --
# the filename's letter is the signal, checked before the bare-UC lookup below.
owner_of = {}
for i in impl:
    for uc in (i.get('traces_to') or []):
        prev = owner_of.get(uc)
        if prev is None or (_is_primary(i['id']) and not _is_primary(prev)):
            owner_of[uc] = i['id']
lock = {}
if os.path.exists(RENDER_LOCK):
    lock = yaml.safe_load(open(RENDER_LOCK, encoding='utf-8')) or {}

stale, missing = [], []
for src in sorted(glob.glob('docs/diagrams/seq-uc*.drawio')):
    src = src.replace(os.sep, '/')
    name = os.path.basename(src)
    m = re.match(r'seq-uc(\d+)([a-z]?)-', name)
    num, suffix = m.group(1), m.group(2)
    if suffix:
        prefix = 'UC{}{}'.format(num, suffix.upper())
        issue = next((i['id'] for i in impl if i['id'].startswith(prefix)), None)
        if not issue:
            F('{} names a suffixed issue id starting {} but no tracker issue matches'
              .format(name, prefix))
            continue
    else:
        issue = owner_of.get('UC-' + num)
        if not issue:
            F('{} draws UC-{} which no implementation issue traces to'.format(name, num))
            continue
    png = 'pm/issues/{}/{}'.format(_slugdir(issue), name[:-7] + '.png')
    if not os.path.exists(png):
        missing.append('{} -> {}'.format(name, png))
        continue
    cur = _sha(src)
    was = (lock.get(src) or {}).get('source_sha')
    if was is None:
        missing.append('{} is not in {}'.format(name, RENDER_LOCK))
    elif was != cur:
        stale.append(name)
for f in sorted(glob.glob('docs/diagrams/seq-uc*.png')):
    F('sequence render left in docs/diagrams/ -- it belongs in its issue folder and this '
      'copy will go stale: {}'.format(os.path.basename(f)))
for x in missing:
    F('sequence render missing or unrecorded: {}'.format(x))
if stale:
    F('.drawio changed since its render was exported, so the committed picture is wrong: '
      '{} -- re-export, then `python audit.py --record-renders`'.format(', '.join(stale)))
if not (stale or missing):
    O('{} sequence renders present in their issue folder and current with their source'
      .format(len(lock)))

# ---------- what is still open, reported not judged ----------
sec4 = s.split('## 4. Not decided')[1].split('###')[0]
still = [r.split('|')[1].strip() for r in sec4.split('\n')
         if r.startswith('|') and 'Item' not in r and '---' not in r]
O('fr-nfr.md section 4 open items: {}'.format(len(still)))
for x in still:
    O('    - {}'.format(x[:96]))

print('=' * 78)
for x in ok:
    print('  PASS  ', x)
for x in warn:
    print('  WARN  ', x)
for x in fail:
    print('  FAIL  ', x)
print('=' * 78)
print('{} passed, {} warnings, {} failures'.format(len(ok), len(warn), len(fail)))
sys.exit(1 if fail else 0)
